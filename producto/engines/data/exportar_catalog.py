"""Exporta el catálogo del motor de scoring a archivos auditables (CSV / XLSX).

Cierra T003 en su forma invertida: la fuente de verdad de las reglas es
``producto/engines/scoring_engine/scoring_engine/catalog.py`` (Python versionado,
revisable en PRs); este script produce vistas legibles para que Melissa y Caro
validen los pesos, categorías y racionales sin leer código.

Uso
---
Desde la raíz del repo::

    cd producto/engines
    python -m data.exportar_catalog                 # solo CSVs (stdlib)
    python -m data.exportar_catalog --xlsx          # además un Excel con hojas (requiere openpyxl)
    python -m data.exportar_catalog --out ./exports # cambia el directorio de salida

Genera en el directorio de salida (por defecto ``producto/engines/data/``):

- ``productos.csv``     — 12 productos: id, nombre, familia
- ``variables.csv``     — 11 variables: código, etiqueta, origen (sistema/formulario), categorías
- ``matriz_pesos.csv``  — 32 filas x 12 productos; cada celda es el peso 0..5 aportado por (variable, categoría) a cada producto, más el racional
- ``triggers.csv``      — disparadores duros (hecho verificable → producto entra al top-N)
- ``checklist.csv``     — datos necesarios para cerrar cada producto y su modo (con/sin intermediario)
- ``fuentes.csv``       — fuentes públicas que sustentan los pesos
- ``catalog_export.xlsx`` (solo con ``--xlsx``) — un único Excel con todas las hojas anteriores

Los archivos son SOLO LECTURA para el motor: cualquier cambio pedido se aplica
editando ``catalog.py`` y re-ejecutando este export.
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path


def _project_root() -> Path:
    """Ubica la raíz del proyecto ``producto/engines/`` para armar sys.path."""
    return Path(__file__).resolve().parent.parent


def _prepare_import_path() -> None:
    sys.path.insert(0, str(_project_root() / "scoring_engine"))


def _write_csv(path: Path, header: list[str], rows: list[list]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        writer.writerows(rows)


def _productos_rows(PRODUCTS) -> tuple[list[str], list[list]]:
    header = ["key", "nombre", "linea"]
    rows = [[p.key, p.nombre, p.linea] for p in PRODUCTS]
    return header, rows


def _variables_rows(VARIABLES) -> tuple[list[str], list[list]]:
    header = ["code", "label", "origen", "categoria"]
    rows = []
    for v in VARIABLES:
        for cat in v.categorias:
            rows.append([v.code, v.label, v.origen, cat])
    return header, rows


def _matriz_rows(PRODUCTS, VARIABLES, WEIGHTS, RATIONALE) -> tuple[list[str], list[list]]:
    prod_keys = [p.key for p in PRODUCTS]
    header = ["variable", "etiqueta_variable", "categoria", *prod_keys, "racional"]
    rows = []
    var_by_code = {v.code: v for v in VARIABLES}
    for v in VARIABLES:
        for cat in v.categorias:
            key = f"{v.code}|{cat}"
            pesos = WEIGHTS.get(key)
            if pesos is None:
                continue
            racional = RATIONALE.get(key, "")
            rows.append([v.code, var_by_code[v.code].label, cat, *pesos, racional])
    return header, rows


def _triggers_rows(TRIGGERS) -> tuple[list[str], list[list]]:
    header = ["variable", "categoria", "producto_key", "motivo"]
    rows = [[t.code, t.categoria, t.product_key, t.motivo] for t in TRIGGERS]
    return header, rows


def _checklist_rows(CHECKLIST) -> tuple[list[str], list[list]]:
    header = ["producto_key", "modo", "item"]
    rows = []
    for prod_key, chk in CHECKLIST.items():
        items = chk.items if chk.items else ("(sin ítems declarados)",)
        for item in items:
            rows.append([prod_key, chk.modo, item])
    return header, rows


def _fuentes_rows(SOURCES) -> tuple[list[str], list[list]]:
    header = ["tema", "cifra", "fuente"]
    rows = [list(row) for row in SOURCES]
    return header, rows


def _write_xlsx(out_dir: Path, sheets: dict[str, tuple[list[str], list[list]]]) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    heat = {
        0: "FFFFFF", 1: "E8F0FE", 2: "C7DBFC",
        3: "9FC1F9", 4: "6FA3F5", 5: "3F86F0",
    }

    for sheet_name, (header, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)

        if sheet_name == "matriz_pesos":
            peso_cols = range(4, 4 + len(header) - 4 - 1)
            for r_idx in range(2, ws.max_row + 1):
                for c_idx in peso_cols:
                    cell = ws.cell(row=r_idx, column=c_idx)
                    color = heat.get(cell.value, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.alignment = Alignment(horizontal="center")

        for col_idx, col_name in enumerate(header, start=1):
            max_len = max(
                [len(str(col_name))]
                + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 80)
        ws.freeze_panes = "A2"

    path = out_dir / "catalog_export.xlsx"
    wb.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Exporta scoring_engine.catalog a CSV / XLSX auditables.")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Directorio de salida (por defecto: junto a este script).",
    )
    parser.add_argument("--xlsx", action="store_true", help="Además de CSVs, genera un Excel con hojas.")
    args = parser.parse_args(argv)

    _prepare_import_path()
    from scoring_engine.catalog import (
        CHECKLIST,
        PRODUCTS,
        RATIONALE,
        SOURCES,
        TRIGGERS,
        VARIABLES,
        WEIGHTS,
    )

    out_dir: Path = args.out
    out_dir.mkdir(parents=True, exist_ok=True)

    sheets: dict[str, tuple[list[str], list[list]]] = {
        "productos": _productos_rows(PRODUCTS),
        "variables": _variables_rows(VARIABLES),
        "matriz_pesos": _matriz_rows(PRODUCTS, VARIABLES, WEIGHTS, RATIONALE),
        "triggers": _triggers_rows(TRIGGERS),
        "checklist": _checklist_rows(CHECKLIST),
        "fuentes": _fuentes_rows(SOURCES),
    }

    for name, (header, rows) in sheets.items():
        path = out_dir / f"{name}.csv"
        _write_csv(path, header, rows)
        print(f"  ✓ {path.relative_to(_project_root().parent)}  ({len(rows)} filas)")

    if args.xlsx:
        xlsx_path = _write_xlsx(out_dir, sheets)
        if xlsx_path is None:
            print("  ⚠ openpyxl no está instalado — omito XLSX. Instala con: pip install openpyxl")
            return 1
        print(f"  ✓ {xlsx_path.relative_to(_project_root().parent)}  (heat-map en matriz_pesos)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
